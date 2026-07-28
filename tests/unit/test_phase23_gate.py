"""Phase 23 integration gate (WP-23.6).

Verifies Phase 23 features work end-to-end across module boundaries.
No live Qdrant/Ollama required; no frontend assertions (WP-23.1 scroll is
manual browser verification).

Features exercised:
  WP-23.2 — XLSX export via checklist service → to_xlsx()
  WP-23.3 — quality warnings wired into chunk_text (overlap guard, contiguity)
  WP-23.4 — cybersecurity profile has skip_sections; param accepted by both
             chunk_text entry points
  WP-23.5 — recovered_truncated flows from extraction through parse_and_normalize
  Regression — CSV/JSON/MD checklist generation unchanged by Phase 23
"""
import io
import json
import shutil
from pathlib import Path

import openpyxl
import pytest

from core.profiles import load_profile
from pipeline.checklist_export import to_csv, to_json, to_markdown, to_xlsx
from pipeline.chunk_text import chunk_text, validate_page_contiguity
from pipeline.llm_extract_requirements import extract_json_array
from pipeline.parse_and_normalize import run as normalize_run
from services.checklist_service import generate

_FIXTURE = Path(__file__).resolve().parent.parent / "fixtures" / "sample_normalized_reqs.jsonl"
DOC_KEY = "SYNTHETIC_TEST_DOC"


@pytest.fixture()
def processed_dir(tmp_path: Path) -> Path:
    run_dir = tmp_path / f"{DOC_KEY}_20260101_000000"
    run_dir.mkdir()
    shutil.copy(_FIXTURE, run_dir / f"{DOC_KEY}_requirements_normalized.jsonl")
    return tmp_path


# ---------------------------------------------------------------------------
# WP-23.2 — XLSX export integration (service → workbook)
# ---------------------------------------------------------------------------

def test_xlsx_from_service_produces_valid_workbook(processed_dir):
    checklist = generate(processed_dir, DOC_KEY, "cybersecurity")
    raw = to_xlsx(checklist)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert wb is not None


def test_xlsx_item_count_matches_checklist(processed_dir):
    checklist = generate(processed_dir, DOC_KEY, "cybersecurity")
    raw = to_xlsx(checklist)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active
    data_rows = ws.max_row - 2  # two header rows
    assert data_rows == len(checklist["items"])


def test_xlsx_has_frozen_pane(processed_dir):
    checklist = generate(processed_dir, DOC_KEY, "cybersecurity")
    raw = to_xlsx(checklist)
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    ws = wb.active
    assert ws.freeze_panes is not None


def test_xlsx_api_route_format_accepted(processed_dir):
    # Smoke: the 'xlsx' format string is wired into the route layer.
    # Import the route module to confirm it doesn't error on import.
    from api.routes.checklist import router  # noqa: F401


# ---------------------------------------------------------------------------
# WP-23.3 — quality warning functions wired correctly
# ---------------------------------------------------------------------------

def test_overlap_guard_raises_on_equal_size():
    with pytest.raises(ValueError, match="overlap"):
        chunk_text("text content here", [(0, 16, 1)], chunk_size=10, overlap=10)


def test_contiguity_validates_without_raising():
    # validate_page_contiguity logs warnings for gaps/duplicates but must not raise.
    # Warning content is verified in test_chunk_text.py.
    pages = [{"page_num": 1, "text": "a"}, {"page_num": 3, "text": "b"}]
    validate_page_contiguity(pages)


def test_overlap_guard_forward_progress_caught():
    with pytest.raises(ValueError, match="forward progress"):
        chunk_text("hello world foo bar", [(0, 19, 1)], chunk_size=10, overlap=9)


# ---------------------------------------------------------------------------
# WP-23.4 — skip_sections in cybersecurity profile + param threading
# ---------------------------------------------------------------------------

def test_cybersecurity_profile_has_skip_sections():
    profile = load_profile("cybersecurity")
    assert "skip_sections" in profile
    skip = profile["skip_sections"]
    assert isinstance(skip, list)
    assert len(skip) > 0


def test_cybersecurity_profile_skip_sections_contains_expected_entries():
    profile = load_profile("cybersecurity")
    skip_upper = [s.upper() for s in profile["skip_sections"]]
    for expected in ("GLOSSARY", "REFERENCES", "ACRONYMS"):
        assert expected in skip_upper, f"Expected '{expected}' in skip_sections"


def test_chunk_text_run_accepts_skip_sections_param():
    # Legacy run() accepts skip_sections without raising
    import inspect

    import pipeline.chunk_text as ct
    sig = inspect.signature(ct.run)
    assert "skip_sections" in sig.parameters


def test_chunk_text_run_structure_aware_accepts_skip_sections_param():
    import inspect

    import pipeline.chunk_text as ct
    sig = inspect.signature(ct.run_structure_aware)
    assert "skip_sections" in sig.parameters


# ---------------------------------------------------------------------------
# WP-23.5 — recovered_truncated flows through parse_and_normalize
# ---------------------------------------------------------------------------

def _write_jsonl(path: Path, records: list[dict]) -> None:
    with open(path, "w") as f:
        for rec in records:
            f.write(json.dumps(rec) + "\n")


@pytest.fixture()
def normalize_dir(tmp_path: Path) -> Path:
    return tmp_path


def test_recovered_truncated_true_preserved_in_normalized(normalize_dir):
    reqs_path = normalize_dir / "doc_extracted_requirements.jsonl"
    chunks_path = normalize_dir / "doc_chunks.jsonl"
    _write_jsonl(reqs_path, [{
        "chunk_id": 1,
        "requirement_id": "R-1-0",
        "source_quote": "Systems shall encrypt data at rest.",
        "source_ref": "3.1",
        "description": "",
        "domain_tags": [],
        "requirement_type": "",
        "recovered_truncated": True,
    }])
    _write_jsonl(chunks_path, [{"chunk_id": 1, "page_start": 1, "page_end": 1, "text": "Systems shall encrypt data at rest."}])
    out_path = normalize_run(str(reqs_path), str(chunks_path), "", str(normalize_dir))
    records = [json.loads(line) for line in Path(out_path).read_text().splitlines() if line.strip()]
    assert len(records) == 1
    assert records[0].get("recovered_truncated") is True


def test_recovered_truncated_false_for_normal_extraction(normalize_dir):
    reqs_path = normalize_dir / "doc2_extracted_requirements.jsonl"
    chunks_path = normalize_dir / "doc2_chunks.jsonl"
    _write_jsonl(reqs_path, [{
        "chunk_id": 1,
        "requirement_id": "R-1-0",
        "source_quote": "Organizations must implement access controls.",
        "source_ref": "AC-1",
        "description": "",
        "domain_tags": [],
        "requirement_type": "",
    }])
    _write_jsonl(chunks_path, [{"chunk_id": 1, "page_start": 1, "page_end": 1, "text": "Organizations must implement access controls."}])
    out_path = normalize_run(str(reqs_path), str(chunks_path), "", str(normalize_dir))
    records = [json.loads(line) for line in Path(out_path).read_text().splitlines() if line.strip()]
    assert len(records) == 1
    assert records[0].get("recovered_truncated") is False


def test_extract_json_array_returns_tuple():
    result, recovered = extract_json_array('[{"source_quote": "test", "source_ref": "1"}]')
    assert isinstance(result, list)
    assert isinstance(recovered, bool)
    assert recovered is False


def test_extract_json_array_truncated_returns_true():
    raw = '[{"source_quote": "test", "source_ref": "1"}, {"source_quote": "inc'
    result, recovered = extract_json_array(raw)
    assert result is not None
    assert recovered is True


# ---------------------------------------------------------------------------
# Regression — Phase 21/22 checklist formats unchanged
# ---------------------------------------------------------------------------

def test_csv_generation_regression(processed_dir):
    checklist = generate(processed_dir, DOC_KEY, "cybersecurity")
    result = to_csv(checklist)
    assert isinstance(result, str)
    assert "source_quote" in result


def test_json_generation_regression(processed_dir):
    checklist = generate(processed_dir, DOC_KEY, "cybersecurity")
    result = to_json(checklist)
    data = json.loads(result)
    assert "items" in data
    assert len(data["items"]) > 0


def test_markdown_generation_regression(processed_dir):
    checklist = generate(processed_dir, DOC_KEY, "cybersecurity")
    result = to_markdown(checklist)
    assert isinstance(result, str)
    assert "#" in result


def test_all_four_export_formats_produce_output(processed_dir):
    checklist = generate(processed_dir, DOC_KEY, "cybersecurity")
    assert to_csv(checklist)
    assert to_json(checklist)
    assert to_markdown(checklist)
    assert to_xlsx(checklist)

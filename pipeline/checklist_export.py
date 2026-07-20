"""Checklist export — serialize a checklist envelope dict to CSV, JSON, Markdown, or XLSX.

Input:  the dict returned by services/checklist_service.py::generate()
Output: a string (CSV/JSON/Markdown) or bytes (XLSX); caller writes to file or stdout.

All display and export logic lives here and in cli/reqbot.py (WP-21.5).
"""
import csv
import io
import json

# CSV column order: locate → ask → record → verify → trace
_CSV_COLUMNS = [
    "source_ref",
    "section_title_path",
    "page_refs",
    "source_quote",
    "audit_question",
    "status",
    "assessor_notes",
    "requires_human_review",
    "review_reasons",
    "confidence",
    "checklist_item_id",
    "requirement_ids",
    "domain_tags",
]


_FORMULA_CHARS = frozenset("=+-@")


def _join(values: list, sep: str) -> str:
    return sep.join(str(v) for v in values)


def _csv_safe(value: object) -> object:
    """Prefix formula-like string cells with a single quote to block spreadsheet injection.

    Excel and LibreOffice treat cells whose effective first character is =, +, -, or @
    as formulas regardless of CSV quoting. Prefixing with ' is the standard mitigation.
    Non-string values (bool, float) are returned unchanged.
    """
    if isinstance(value, str):
        stripped = value.lstrip()
        if stripped and stripped[0] in _FORMULA_CHARS:
            return "'" + value
    return value


def _csv_row(item: dict) -> dict:
    raw = {
        "source_ref": item.get("source_ref", ""),
        "section_title_path": _join(item.get("section_title_path") or [], " > "),
        "page_refs": _join(item.get("page_refs") or [], ", "),
        "source_quote": item.get("source_quote", ""),
        "audit_question": item.get("audit_question", ""),
        "status": item.get("status", ""),
        "assessor_notes": item.get("assessor_notes", ""),
        "requires_human_review": item.get("requires_human_review", False),
        "review_reasons": _join(item.get("review_reasons") or [], "; "),
        "confidence": item.get("confidence", 0.0),
        "checklist_item_id": item.get("checklist_item_id", ""),
        "requirement_ids": _join(item.get("requirement_ids") or [], ", "),
        "domain_tags": _join(item.get("domain_tags") or [], ", "),
    }
    return {k: _csv_safe(v) for k, v in raw.items()}


def to_csv(checklist: dict) -> str:
    """Return the checklist items as a CSV string (UTF-8, Excel-friendly)."""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_CSV_COLUMNS, lineterminator="\r\n")
    writer.writeheader()
    for item in checklist.get("items", []):
        writer.writerow(_csv_row(item))
    return buf.getvalue()


def to_json(checklist: dict) -> str:
    """Return the full checklist envelope as a pretty-printed JSON string."""
    return json.dumps(checklist, indent=2, default=str)


def _md_item(item: dict, index: int) -> str:
    section = _join(item.get("section_title_path") or [], " > ")
    pages = _join(item.get("page_refs") or [], ", ")
    header_parts = []
    if section:
        header_parts.append(section)
    if pages:
        header_parts.append(f"p. {pages}")
    heading = " — ".join(header_parts) if header_parts else f"Item {index}"

    lines = [f"## {heading}", ""]

    source_ref = item.get("source_ref", "")
    if source_ref:
        lines.append(f"**Source Ref:** {source_ref}  ")
    tags = _join(item.get("domain_tags") or [], ", ")
    if tags:
        lines.append(f"**Domain Tags:** {tags}  ")
    conf = item.get("confidence")
    if conf is None:
        conf = 0.0
    lines.append(f"**Confidence:** {conf:.2f}  ")
    lines.append("")

    quote = item.get("source_quote", "")
    lines.append(f"> {quote}")
    lines.append("")

    audit_q = item.get("audit_question", "")
    lines.append(f"**Audit Question:** {audit_q if audit_q else '*(not generated)*'}  ")

    status = item.get("status", "not-started")
    assessor = item.get("assessor_notes", "")
    lines.append(f"**Status:** {status}  ")
    lines.append(f"**Assessor Notes:** {assessor if assessor else '*(none)*'}  ")
    lines.append("")

    if item.get("requires_human_review"):
        reasons = _join(item.get("review_reasons") or [], "; ")
        lines.append(f"> ⚠ **Requires Review:** {reasons}  ")
        lines.append("")

    req_ids = _join(item.get("requirement_ids") or [], ", ")
    lines.append(f"*ID: {item.get('checklist_item_id', '')} | Req: {req_ids}*")
    lines.append("")
    lines.append("---")
    lines.append("")
    return "\n".join(lines)


def to_xlsx(checklist: dict) -> bytes:
    """Return the checklist items as an Excel workbook (XLSX) in bytes.

    Requires openpyxl (approved WP-23.2 dependency).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    # Column definitions: (display_header, item_key, col_width, wrap_text)
    _COLS = [
        ("Ref",            "source_ref",            14, False),
        ("Section",        "section_title_path",     28, False),
        ("Pages",          "page_refs",               9, False),
        ("Source Quote",   "source_quote",            36, True),
        ("Audit Question", "audit_question",          22, True),
        ("Status",         "status",                  15, False),
        ("Notes",          "assessor_notes",          22, True),
        ("Flag",           "requires_human_review",    8, False),
        ("Reasons",        "review_reasons",          22, False),
        ("Conf.",          "confidence",               7, False),
        ("Item ID",        "checklist_item_id",       26, False),
        ("Req IDs",        "requirement_ids",         20, False),
        ("Tags",           "domain_tags",             20, False),
    ]

    # Column groups: (label, first_col_1based, last_col_1based)
    _GROUPS = [
        ("Locate",  1,  3),
        ("Ask",     4,  5),
        ("Record",  6,  7),
        ("Verify",  8, 10),
        ("Trace",  11, 13),
    ]

    _GROUP_FILL = PatternFill("solid", fgColor="E2E8F0")
    _FLAGGED_FILL = PatternFill("solid", fgColor="FFFBEB")
    _HEADER_FONT = Font(bold=True, size=9)

    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"

    # Row 1: group header row
    for label, c_start, c_end in _GROUPS:
        cell = ws.cell(row=1, column=c_start, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _GROUP_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if c_start != c_end:
            ws.merge_cells(
                start_row=1, start_column=c_start,
                end_row=1, end_column=c_end,
            )

    # Row 2: column header row
    for col_idx, (header, _, _, _) in enumerate(_COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _GROUP_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Freeze panes below both header rows, after Locate group (3 cols)
    ws.freeze_panes = "D3"

    # Auto-filter anchored to column header row
    ws.auto_filter.ref = f"A2:{get_column_letter(len(_COLS))}2"

    # Status column data validation (dropdown)
    dv = DataValidation(
        type="list",
        formula1='"not-started,in-progress,compliant,non-compliant,not-applicable"',
        allow_blank=True,
    )
    ws.add_data_validation(dv)

    # Column widths
    for col_idx, (_, _, width, _) in enumerate(_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Data rows
    for item in checklist.get("items", []):
        flagged = item.get("requires_human_review", False)
        row_fill = _FLAGGED_FILL if flagged else None
        confidence_val = item.get("confidence") or 0.0

        values = [
            _csv_safe(item.get("source_ref") or ""),
            _csv_safe(_join(item.get("section_title_path") or [], " > ")),
            _csv_safe(_join(item.get("page_refs") or [], ", ")),
            _csv_safe(item.get("source_quote") or ""),
            _csv_safe(item.get("audit_question") or ""),
            _csv_safe(item.get("status") or ""),
            _csv_safe(item.get("assessor_notes") or ""),
            "Yes" if flagged else "No",
            _csv_safe(_join(item.get("review_reasons") or [], "; ")),
            confidence_val,
            _csv_safe(item.get("checklist_item_id") or ""),
            _csv_safe(_join(item.get("requirement_ids") or [], ", ")),
            _csv_safe(_join(item.get("domain_tags") or [], ", ")),
        ]

        row_num = ws.max_row + 1
        for col_idx, (_, _, _, wrap) in enumerate(_COLS, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=values[col_idx - 1])
            cell.alignment = Alignment(vertical="top", wrap_text=wrap)
            if row_fill:
                cell.fill = row_fill

        # Confidence as percentage (value is 0–1 float)
        ws.cell(row=row_num, column=10).number_format = "0%"
        # Status: register with data validation
        dv.add(ws.cell(row=row_num, column=6))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_markdown(checklist: dict) -> str:
    """Return the checklist as a Markdown string."""
    doc = checklist.get("document", {})
    source_pdf = doc.get("source_pdf", "")
    title = source_pdf if source_pdf else doc.get("document_id", "Unknown Document")

    summary = checklist.get("summary", {})
    total = summary.get("total_items", 0)
    review_count = summary.get("items_requiring_review", 0)

    lines = [
        f"# Checklist: {title}",
        "",
        f"**Profile:** {checklist.get('profile', '')}  ",
        f"**Generated:** {checklist.get('generated_at', '')}  ",
        f"**Items:** {total} total, {review_count} requiring review  ",
        "",
        "---",
        "",
    ]

    for i, item in enumerate(checklist.get("items", []), start=1):
        lines.append(_md_item(item, i))

    return "\n".join(lines)

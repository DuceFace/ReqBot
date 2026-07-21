import csv
import io
import json

import openpyxl

from pipeline.checklist_export import (
    _CSV_COLUMNS,
    _csv_row,
    _csv_safe,
    to_csv,
    to_json,
    to_markdown,
    to_xlsx,
)

# Minimal fully-populated checklist item for fixture use
COMPLETE_ITEM = {
    "checklist_item_id": "CHK-abcdef1234567890",
    "requirement_ids": ["REQ-abc123"],
    "domain_tags": ["access-control", "authentication"],
    "source_ref": "4.2.1",
    "page_refs": [12, 13],
    "section_title_path": ["Access Control", "MFA Policy"],
    "source_quote": "The system shall enforce MFA for all privileged accounts.",
    "audit_question": "",
    "evidence_to_request": [],
    "generation_notes": "",
    "assessor_notes": "",
    "status": "not-started",
    "confidence": 0.9,
    "requires_human_review": False,
    "review_reasons": [],
}

REVIEW_ITEM = {
    **COMPLETE_ITEM,
    "checklist_item_id": "CHK-bbbbbbbbbbbbbbbb",
    "requirement_ids": ["REQ-def456"],
    "source_ref": "",
    "section_title_path": [],
    "page_refs": [],
    "domain_tags": [],
    "confidence": 0.6,
    "requires_human_review": True,
    "review_reasons": ["missing-source-ref", "low-confidence", "missing-page-refs", "missing-section-title-path", "missing-domain-tags"],
}

ENVELOPE = {
    "format": "reqbot-checklist",
    "format_version": "1.0",
    "generated_at": "2026-07-19T12:00:00+00:00",
    "generator": {"tool": "reqbot", "command": "reqbot checklist --doc test-doc --profile cybersecurity"},
    "document": {"document_id": "abc123", "source_pdf": "test-doc.pdf"},
    "profile": "cybersecurity",
    "summary": {"total_items": 2, "items_requiring_review": 1},
    "items": [COMPLETE_ITEM, REVIEW_ITEM],
}

EMPTY_ENVELOPE = {**ENVELOPE, "summary": {"total_items": 0, "items_requiring_review": 0}, "items": []}


# --- _csv_row ---

def test_csv_row_array_joins():
    row = _csv_row(COMPLETE_ITEM)
    assert row["section_title_path"] == "Access Control > MFA Policy"
    assert row["page_refs"] == "12, 13"
    assert row["requirement_ids"] == "REQ-abc123"
    assert row["domain_tags"] == "access-control, authentication"


def test_csv_row_review_reasons_semicolon_join():
    row = _csv_row(REVIEW_ITEM)
    assert "missing-source-ref" in row["review_reasons"]
    assert ";" in row["review_reasons"]


def test_csv_row_empty_arrays_produce_empty_strings():
    row = _csv_row(REVIEW_ITEM)
    assert row["section_title_path"] == ""
    assert row["page_refs"] == ""
    assert row["domain_tags"] == ""
    assert row["review_reasons"].count(";") >= 1  # multiple reasons joined


def test_csv_row_none_arrays_produce_empty_strings():
    item = {**COMPLETE_ITEM, "section_title_path": None, "page_refs": None, "domain_tags": None, "review_reasons": None}
    row = _csv_row(item)
    assert row["section_title_path"] == ""
    assert row["page_refs"] == ""
    assert row["domain_tags"] == ""
    assert row["review_reasons"] == ""


# --- to_csv ---

def test_to_csv_returns_string():
    result = to_csv(ENVELOPE)
    assert isinstance(result, str)


def test_to_csv_column_order():
    result = to_csv(ENVELOPE)
    reader = csv.reader(io.StringIO(result))
    header = next(reader)
    assert header == _CSV_COLUMNS


def test_to_csv_row_count():
    result = to_csv(ENVELOPE)
    reader = csv.reader(io.StringIO(result))
    rows = list(reader)
    assert len(rows) == 3  # header + 2 items


def test_to_csv_empty_items():
    result = to_csv(EMPTY_ENVELOPE)
    reader = csv.reader(io.StringIO(result))
    rows = list(reader)
    assert len(rows) == 1  # header only


def test_to_csv_source_quote_present():
    result = to_csv(ENVELOPE)
    assert "The system shall enforce MFA" in result


def test_to_csv_requires_human_review_serialized():
    result = to_csv(ENVELOPE)
    assert "False" in result
    assert "True" in result


def test_to_csv_confidence_in_output():
    result = to_csv(ENVELOPE)
    assert "0.9" in result


def test_to_csv_no_crash_on_missing_item_fields():
    sparse_item = {"requirement_ids": ["REQ-001"], "source_quote": "Minimal."}
    checklist = {**EMPTY_ENVELOPE, "items": [sparse_item]}
    result = to_csv(checklist)
    reader = csv.reader(io.StringIO(result))
    rows = list(reader)
    assert len(rows) == 2


# --- to_json ---

def test_to_json_returns_valid_json():
    result = to_json(ENVELOPE)
    parsed = json.loads(result)
    assert parsed["format"] == "reqbot-checklist"


def test_to_json_roundtrip_preserves_items():
    result = to_json(ENVELOPE)
    parsed = json.loads(result)
    assert len(parsed["items"]) == 2


def test_to_json_is_pretty_printed():
    result = to_json(ENVELOPE)
    assert "\n" in result


def test_to_json_empty_items():
    result = to_json(EMPTY_ENVELOPE)
    parsed = json.loads(result)
    assert parsed["items"] == []


def test_to_json_preserves_envelope_fields():
    result = to_json(ENVELOPE)
    parsed = json.loads(result)
    assert parsed["profile"] == "cybersecurity"
    assert parsed["document"]["source_pdf"] == "test-doc.pdf"


# --- to_markdown ---

def test_to_markdown_returns_string():
    result = to_markdown(ENVELOPE)
    assert isinstance(result, str)


def test_to_markdown_title_includes_source_pdf():
    result = to_markdown(ENVELOPE)
    assert "test-doc.pdf" in result


def test_to_markdown_profile_present():
    result = to_markdown(ENVELOPE)
    assert "cybersecurity" in result


def test_to_markdown_source_quote_present():
    result = to_markdown(ENVELOPE)
    assert "The system shall enforce MFA" in result


def test_to_markdown_section_title_path_in_heading():
    result = to_markdown(ENVELOPE)
    assert "Access Control" in result
    assert "MFA Policy" in result


def test_to_markdown_page_refs_in_heading():
    result = to_markdown(ENVELOPE)
    assert "p. 12" in result


def test_to_markdown_review_flag_shown_for_flagged_item():
    result = to_markdown(ENVELOPE)
    assert "Requires Review" in result
    assert "low-confidence" in result


def test_to_markdown_no_review_flag_for_clean_item():
    result = to_markdown({**EMPTY_ENVELOPE, "items": [COMPLETE_ITEM]})
    # Verify the item itself rendered (not just an empty list)
    assert "The system shall enforce MFA" in result
    assert "Requires Review" not in result


def test_to_markdown_audit_question_blank_shows_not_generated():
    result = to_markdown(ENVELOPE)
    assert "*(not generated)*" in result


def test_to_markdown_audit_question_shown_when_present():
    item = {**COMPLETE_ITEM, "audit_question": "Has MFA been enforced?"}
    checklist = {**EMPTY_ENVELOPE, "items": [item]}
    result = to_markdown(checklist)
    assert "Has MFA been enforced?" in result


def test_to_markdown_empty_items_no_crash():
    result = to_markdown(EMPTY_ENVELOPE)
    assert "test-doc.pdf" in result
    assert "0 total" in result


def test_to_markdown_item_id_and_req_id_in_output():
    result = to_markdown(ENVELOPE)
    assert "CHK-abcdef1234567890" in result
    assert "REQ-abc123" in result


def test_to_markdown_no_source_pdf_falls_back_to_document_id():
    checklist = {**EMPTY_ENVELOPE, "document": {"document_id": "abc123", "source_pdf": ""}}
    result = to_markdown(checklist)
    assert "abc123" in result


def test_to_markdown_null_confidence_renders_as_zero():
    item = {**COMPLETE_ITEM, "confidence": None}
    checklist = {**EMPTY_ENVELOPE, "items": [item]}
    result = to_markdown(checklist)
    assert "0.00" in result


# --- _csv_safe (formula injection prevention) ---

def test_csv_safe_equals_prefix_escaped():
    assert _csv_safe("=SUM(A1:A10)") == "'=SUM(A1:A10)"


def test_csv_safe_plus_prefix_escaped():
    assert _csv_safe("+1234") == "'+1234"


def test_csv_safe_minus_prefix_escaped():
    assert _csv_safe("-DROP TABLE") == "'-DROP TABLE"


def test_csv_safe_at_prefix_escaped():
    assert _csv_safe("@user") == "'@user"


def test_csv_safe_whitespace_then_formula_escaped():
    assert _csv_safe("  =dangerous") == "'  =dangerous"


def test_csv_safe_tab_then_formula_escaped():
    assert _csv_safe("\t=foo") == "'\t=foo"


def test_csv_safe_normal_text_unchanged():
    assert _csv_safe("Systems must enforce MFA.") == "Systems must enforce MFA."


def test_csv_safe_empty_string_unchanged():
    assert _csv_safe("") == ""


def test_csv_safe_whitespace_only_unchanged():
    assert _csv_safe("   ") == "   "


def test_csv_safe_nonstring_bool_unchanged():
    assert _csv_safe(True) is True


def test_csv_safe_nonstring_float_unchanged():
    assert _csv_safe(0.9) == 0.9


# --- CSV formula injection integration ---

def test_to_csv_formula_source_quote_escaped():
    item = {**COMPLETE_ITEM, "source_quote": "=HYPERLINK(\"http://evil.example\",\"click\")"}
    checklist = {**EMPTY_ENVELOPE, "items": [item]}
    result = to_csv(checklist)
    assert "'=HYPERLINK" in result


def test_to_csv_formula_source_ref_escaped():
    item = {**COMPLETE_ITEM, "source_ref": "+1.2.3"}
    checklist = {**EMPTY_ENVELOPE, "items": [item]}
    result = to_csv(checklist)
    assert "'+1.2.3" in result


def test_to_json_formula_value_not_escaped():
    """JSON export must not mutate the checklist envelope."""
    item = {**COMPLETE_ITEM, "source_quote": "=dangerous"}
    checklist = {**EMPTY_ENVELOPE, "items": [item]}
    result = to_json(checklist)
    parsed = json.loads(result)
    assert parsed["items"][0]["source_quote"] == "=dangerous"


def test_to_markdown_formula_value_not_escaped():
    """Markdown export must not apply CSV escaping."""
    item = {**COMPLETE_ITEM, "source_quote": "=Systems shall enforce MFA."}
    checklist = {**EMPTY_ENVELOPE, "items": [item]}
    result = to_markdown(checklist)
    assert "=Systems shall enforce MFA." in result
    assert "'=Systems" not in result


# --- to_xlsx ---

def _load_xlsx(data: bytes):
    """Helper: load xlsx bytes as an openpyxl workbook."""
    return openpyxl.load_workbook(io.BytesIO(data))


def test_to_xlsx_returns_bytes():
    result = to_xlsx(ENVELOPE)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_to_xlsx_opens_as_workbook():
    result = to_xlsx(ENVELOPE)
    wb = _load_xlsx(result)
    assert wb is not None


def test_to_xlsx_sheet_name():
    wb = _load_xlsx(to_xlsx(ENVELOPE))
    assert "Checklist" in wb.sheetnames


def test_to_xlsx_group_headers_in_row_1():
    wb = _load_xlsx(to_xlsx(ENVELOPE))
    ws = wb["Checklist"]
    row1_values = [ws.cell(row=1, column=c).value for c in range(1, 14)]
    assert "Locate" in row1_values
    assert "Ask" in row1_values
    assert "Record" in row1_values
    assert "Verify" in row1_values
    assert "Trace" in row1_values


def test_to_xlsx_column_order():
    expected_headers = [
        "Ref", "Section", "Pages",
        "Source Quote", "Audit Question",
        "Status", "Notes",
        "Flag", "Reasons", "Conf.",
        "Item ID", "Req IDs", "Tags",
    ]
    wb = _load_xlsx(to_xlsx(ENVELOPE))
    ws = wb["Checklist"]
    actual = [ws.cell(row=2, column=c).value for c in range(1, 14)]
    assert actual == expected_headers


def test_to_xlsx_freeze_panes():
    wb = _load_xlsx(to_xlsx(ENVELOPE))
    ws = wb["Checklist"]
    assert ws.freeze_panes == "D3"


def test_to_xlsx_auto_filter_covers_data_rows():
    # ENVELOPE has 2 items → 2 header rows + 2 data rows = max_row 4; filter must end at row 4
    wb = _load_xlsx(to_xlsx(ENVELOPE))
    ws = wb["Checklist"]
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.endswith("4"), f"expected filter to row 4, got {ws.auto_filter.ref}"


def test_to_xlsx_auto_filter_empty_checklist_ends_at_row_2():
    wb = _load_xlsx(to_xlsx(EMPTY_ENVELOPE))
    ws = wb["Checklist"]
    assert ws.auto_filter.ref is not None
    assert ws.auto_filter.ref.endswith("2"), f"expected filter to row 2, got {ws.auto_filter.ref}"


def test_to_xlsx_status_data_validation_exists():
    wb = _load_xlsx(to_xlsx(ENVELOPE))
    ws = wb["Checklist"]
    assert len(ws.data_validations.dataValidation) > 0


def test_to_xlsx_row_count():
    wb = _load_xlsx(to_xlsx(ENVELOPE))
    ws = wb["Checklist"]
    # 2 header rows + 2 data rows
    assert ws.max_row == 4


def test_to_xlsx_source_quote_present():
    wb = _load_xlsx(to_xlsx(ENVELOPE))
    ws = wb["Checklist"]
    all_values = [ws.cell(row=r, column=4).value for r in range(3, ws.max_row + 1)]
    assert any("enforce MFA" in str(v) for v in all_values if v)


def test_to_xlsx_empty_items_no_crash():
    result = to_xlsx(EMPTY_ENVELOPE)
    wb = _load_xlsx(result)
    ws = wb["Checklist"]
    assert ws.max_row == 2  # only 2 header rows


def test_to_xlsx_flagged_row_has_fill():
    wb = _load_xlsx(to_xlsx(ENVELOPE))
    ws = wb["Checklist"]
    # REVIEW_ITEM is second data row (row 4); COMPLETE_ITEM is row 3
    flagged_fill = ws.cell(row=4, column=1).fill
    clean_fill = ws.cell(row=3, column=1).fill
    assert flagged_fill.patternType == "solid"
    assert flagged_fill.fgColor.rgb != clean_fill.fgColor.rgb


def test_to_xlsx_confidence_percentage_format():
    wb = _load_xlsx(to_xlsx(ENVELOPE))
    ws = wb["Checklist"]
    conf_cell = ws.cell(row=3, column=10)
    assert conf_cell.number_format == "0%"


def test_to_xlsx_formula_injection_protection():
    item = {**COMPLETE_ITEM, "source_quote": "=HYPERLINK(\"http://evil.example\",\"click\")"}
    checklist = {**EMPTY_ENVELOPE, "items": [item]}
    wb = _load_xlsx(to_xlsx(checklist))
    ws = wb["Checklist"]
    value = ws.cell(row=3, column=4).value
    assert value is not None
    assert str(value).startswith("'")


def test_to_xlsx_sparse_item_no_crash():
    sparse = {"requirement_ids": ["REQ-001"], "source_quote": "Minimal requirement."}
    checklist = {**EMPTY_ENVELOPE, "items": [sparse]}
    result = to_xlsx(checklist)
    wb = _load_xlsx(result)
    assert wb["Checklist"].max_row == 3
